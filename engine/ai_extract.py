"""AI(Anthropic Claude / OpenAI ChatGPT / Google Gemini)를 이용해 크롤링한 HTML
페이지에서 사용자가 원하는 정보만 구조화된 형태로 뽑아내고, CSV/JSON으로 내보낸다.

크롤링 방식(HTTrack 정적 다운로드든 스마트 크롤링의 브라우저 렌더링이든)과
무관하게, 저장된 폴더 안의 .html 파일들에 대해 동작하는 공용 후처리 단계로
설계했다 - 그래야 어떤 방식으로 받은 프로젝트든 나중에 AI 추출을 돌릴 수 있다.

세 프로바이더 모두 "구조화 출력 강제"로 받는다 (자유 텍스트 응답을 파싱하지 않음):
- Anthropic: tool-use를 tool_choice로 강제 호출
- OpenAI: function calling을 tool_choice로 강제 호출
- Gemini: response_mime_type='application/json' + response_schema로 강제
필드 스키마를 그대로 각 프로바이더의 스키마 형식으로 변환해서 넘기기 때문에,
어느 프로바이더를 쓰든 매 페이지마다 정확히 같은 키를 가진 결과가 나온다
(그래야 CSV 컬럼이 페이지마다/프로바이더마다 달라지지 않는다).
"""
import os
import re
import csv
import json
import glob
import datetime

PROVIDERS = ('anthropic', 'openai', 'gemini')

DEFAULT_MODELS = {
    'anthropic': 'claude-haiku-4-5-20251001',
    'openai': 'gpt-4o-mini',
    'gemini': 'gemini-2.0-flash',
}

PROVIDER_DISPLAY_NAMES = {
    'anthropic': 'Anthropic (Claude)',
    'openai': 'OpenAI (ChatGPT)',
    'gemini': 'Google (Gemini)',
}

# 페이지가 너무 크면 토큰/비용을 넘지 않도록 앞부분만 사용
_MAX_HTML_CHARS = 60_000

_FIELDS_SCHEMA_DESCRIPTION = '사용자 지시문에 맞춰 웹페이지에서 추출할 필드 목록을 제안한다.'
_EXTRACT_DESCRIPTION = '웹페이지에서 지정된 필드 값을 추출한다. 페이지에 해당 정보가 없으면 빈 문자열/0/false를 사용한다.'


def _strip_html_noise(html):
    """<script>/<style> 태그를 제거해서 토큰을 아낀다 (본문 추출 목적이라 실행 코드는 불필요)."""
    html = re.sub(r'<script\b[^>]*>.*?</script>', '', html, flags=re.DOTALL | re.IGNORECASE)
    html = re.sub(r'<style\b[^>]*>.*?</style>', '', html, flags=re.DOTALL | re.IGNORECASE)
    return html[:_MAX_HTML_CHARS]


def _propose_fields_prompt(instruction, sample_html):
    return (
        f'다음 사용자 지시문에 맞춰 이 웹페이지에서 추출할 필드 목록을 제안해줘.\n\n'
        f'지시문: {instruction}\n\n'
        f'페이지 내용(일부):\n{_strip_html_noise(sample_html)}'
    )


def _extract_prompt(html):
    return f'다음 웹페이지에서 필드 값을 추출해줘.\n\n{_strip_html_noise(html)}'


def _fields_list_schema():
    """propose_fields용 JSON 스키마 (세 프로바이더 모두 JSON Schema 서브셋을 그대로 받아들인다)."""
    return {
        'type': 'object',
        'properties': {
            'fields': {
                'type': 'array',
                'items': {
                    'type': 'object',
                    'properties': {
                        'name': {'type': 'string',
                                  'description': '영문 snake_case 필드 이름 (CSV 컬럼명으로 그대로 쓰임)'},
                        'label': {'type': 'string', 'description': '사람이 보기 좋은 한글 라벨'},
                        'type': {'type': 'string', 'enum': ['string', 'number', 'boolean']},
                    },
                    'required': ['name', 'label', 'type'],
                },
            },
        },
        'required': ['fields'],
    }


def _extract_schema(fields):
    properties = {
        f['name']: {'type': f.get('type', 'string'), 'description': f.get('label', f['name'])}
        for f in fields
    }
    return {
        'type': 'object',
        'properties': properties,
        'required': list(properties.keys()),
    }


# ---------------- Anthropic ----------------

def _anthropic_tool_call(api_key, model, tool_name, description, schema, prompt):
    from anthropic import Anthropic
    client = Anthropic(api_key=api_key)
    tool = {'name': tool_name, 'description': description, 'input_schema': schema}
    message = client.messages.create(
        model=model,
        max_tokens=1536,
        tools=[tool],
        tool_choice={'type': 'tool', 'name': tool_name},
        messages=[{'role': 'user', 'content': prompt}],
    )
    for block in message.content:
        if block.type == 'tool_use' and block.name == tool_name:
            return dict(block.input)
    return None


# ---------------- OpenAI ----------------

def _openai_tool_call(api_key, model, tool_name, description, schema, prompt):
    from openai import OpenAI
    client = OpenAI(api_key=api_key)
    tool = {
        'type': 'function',
        'function': {'name': tool_name, 'description': description, 'parameters': schema},
    }
    response = client.chat.completions.create(
        model=model,
        tools=[tool],
        tool_choice={'type': 'function', 'function': {'name': tool_name}},
        messages=[{'role': 'user', 'content': prompt}],
    )
    tool_calls = response.choices[0].message.tool_calls
    if not tool_calls:
        return None
    return json.loads(tool_calls[0].function.arguments)


# ---------------- Gemini ----------------

def _gemini_json_call(api_key, model, schema, prompt):
    from google import genai
    from google.genai import types
    client = genai.Client(api_key=api_key)
    response = client.models.generate_content(
        model=model,
        contents=prompt,
        config=types.GenerateContentConfig(response_mime_type='application/json', response_schema=schema),
    )
    return json.loads(response.text)


def propose_fields(instruction, sample_html, api_key, provider='anthropic', model=None):
    """자연어 지시문 + 샘플 페이지 하나를 보고 추출할 필드 목록을 제안받는다.
    돌려주는 값: [{'name','label','type'}, ...]"""
    if provider not in PROVIDERS:
        raise ValueError(f'알 수 없는 AI 프로바이더입니다: {provider!r}')
    model = model or DEFAULT_MODELS[provider]
    prompt = _propose_fields_prompt(instruction, sample_html)
    schema = _fields_list_schema()

    if provider == 'anthropic':
        result = _anthropic_tool_call(api_key, model, 'propose_fields', _FIELDS_SCHEMA_DESCRIPTION, schema, prompt)
    elif provider == 'openai':
        result = _openai_tool_call(api_key, model, 'propose_fields', _FIELDS_SCHEMA_DESCRIPTION, schema, prompt)
    elif provider == 'gemini':
        result = _gemini_json_call(api_key, model, schema, prompt)
    else:
        raise ValueError(f'알 수 없는 AI 프로바이더입니다: {provider!r}')

    return (result or {}).get('fields', [])


def extract_fields(html, fields, api_key, provider='anthropic', model=None):
    """확정된 필드 스키마로 페이지 하나에서 값을 추출한다. 실패 시 예외를 던진다(호출부가 페이지 단위로 잡음)."""
    if not fields:
        raise ValueError('추출할 필드가 없습니다.')
    if provider not in PROVIDERS:
        raise ValueError(f'알 수 없는 AI 프로바이더입니다: {provider!r}')
    model = model or DEFAULT_MODELS[provider]
    prompt = _extract_prompt(html)
    schema = _extract_schema(fields)

    if provider == 'anthropic':
        result = _anthropic_tool_call(api_key, model, 'extract', _EXTRACT_DESCRIPTION, schema, prompt)
    elif provider == 'openai':
        result = _openai_tool_call(api_key, model, 'extract', _EXTRACT_DESCRIPTION, schema, prompt)
    elif provider == 'gemini':
        result = _gemini_json_call(api_key, model, schema, prompt)
    else:
        raise ValueError(f'알 수 없는 AI 프로바이더입니다: {provider!r}')

    if result is None:
        raise RuntimeError('AI로부터 추출 결과를 받지 못했습니다.')
    return result


def _extract_list_schema(fields):
    """반복 항목 여러 개를 한 번에 뽑을 때 쓰는 스키마 - 항목 하나짜리 스키마를
    배열로 감싼다. 항목 개수와 무관하게 API 호출은 여전히 1회다."""
    return {
        'type': 'object',
        'properties': {'items': {'type': 'array', 'items': _extract_schema(fields)}},
        'required': ['items'],
    }


def _extract_list_prompt(items_html):
    parts = [f'### 항목 {i + 1} ###\n{_strip_html_noise(h)}' for i, h in enumerate(items_html)]
    return (
        f'아래는 한 페이지 안에 나란히 있던 반복 항목 {len(items_html)}개다. '
        f'각 항목에서 필드 값을 추출해서, 입력 순서와 똑같은 순서로 items 배열에 하나씩 담아줘.\n\n'
        + '\n\n'.join(parts)
    )


def extract_list_fields(items_html, fields, api_key, provider='anthropic', model=None, max_items_per_call=20):
    """반복되는 항목(pattern_detect.detect_repeating_blocks 결과) 여러 개를 한 번에 추출한다.

    항목이 많으면(max_items_per_call 초과) 여러 번에 나눠 부르지만, 그래도
    '항목 하나당 호출 1번'보다는 훨씬 적은 횟수다 (20개씩 묶으면 100개 항목도
    호출 5번). 한 묶음이 실패해도 나머지 묶음은 계속 진행한다."""
    if not fields or not items_html:
        return []
    all_rows = []
    for i in range(0, len(items_html), max_items_per_call):
        chunk = items_html[i:i + max_items_per_call]
        model_ = model or DEFAULT_MODELS[provider]
        schema = _extract_list_schema(fields)
        prompt = _extract_list_prompt(chunk)
        if provider == 'anthropic':
            result = _anthropic_tool_call(api_key, model_, 'extract_list', _EXTRACT_DESCRIPTION, schema, prompt)
        elif provider == 'openai':
            result = _openai_tool_call(api_key, model_, 'extract_list', _EXTRACT_DESCRIPTION, schema, prompt)
        elif provider == 'gemini':
            result = _gemini_json_call(api_key, model_, schema, prompt)
        else:
            raise ValueError(f'알 수 없는 AI 프로바이더입니다: {provider!r}')
        all_rows.extend((result or {}).get('items', []))
    return all_rows


def export_records(records, out_dir, base_name, formats):
    """records: [{field: value, ...}, ...]. formats: ['csv','json'] 중 선택.
    저장된 파일 경로 리스트를 돌려준다."""
    if not records:
        return []
    os.makedirs(out_dir, exist_ok=True)
    saved = []

    if 'csv' in formats:
        csv_path = os.path.join(out_dir, f'{base_name}.csv')
        fieldnames = list(records[0].keys())
        # utf-8-sig: 엑셀에서 한글 CSV를 열었을 때 깨지지 않도록 BOM 포함
        with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for row in records:
                writer.writerow(row)
        saved.append(csv_path)

    if 'json' in formats:
        json_path = os.path.join(out_dir, f'{base_name}.json')
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(records, f, ensure_ascii=False, indent=2)
        saved.append(json_path)

    if 'xlsx' in formats:
        # 사용자 모르게 네트워크에서 패키지를 받아 설치하지 않는다.
        # (exe로 묶은 뒤에는 sys.executable이 앱 자신이라 엉뚱하게 동작하기도 한다.)
        # 엑셀 저장만 건너뛰고 CSV/JSON은 정상적으로 남기는 편이 안전하다.
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            openpyxl = None

    if 'xlsx' in formats and openpyxl is None:
        formats = [f for f in formats if f != 'xlsx']

    if 'xlsx' in formats:

        xlsx_path = os.path.join(out_dir, f'{base_name}.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extracted Data"

        fieldnames = list(records[0].keys())
        ws.append(fieldnames)

        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for row in records:
            ws.append([row.get(f, "") for f in fieldnames])

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width

        wb.save(xlsx_path)
        saved.append(xlsx_path)

    return saved


def find_html_files(folder, max_files=None):
    files = sorted(glob.glob(os.path.join(folder, '**', '*.html'), recursive=True))
    return files[:max_files] if max_files else files


def run_extraction(folder, fields, api_key, log_fn, provider='anthropic', model=None,
                    max_pages=50, export_formats=('csv',)):
    """folder 안의 모든 .html 파일을 순서대로 추출하고 export까지 한 번에 처리하는 헬퍼.
    개별 페이지 실패는 건너뛰고 계속 진행한다 (전체 작업이 죽지 않게).

    페이지 안에 반복되는 항목(상품 카드, 게시글 목록 등)이 있으면 그 페이지
    하나에서 여러 행을 뽑아낸다 - pattern_detect가 AI 호출 없이 구조로 찾아내고,
    실제 추출은 항목 개수와 무관하게 API 호출 1~2회로 끝난다(20개씩 묶어서).
    반복 패턴이 없는 보통의 상세 페이지는 예전처럼 '페이지 1개 = 행 1개'다."""
    import pattern_detect
    html_files = find_html_files(folder, max_pages)
    records = []
    for path in html_files:
        try:
            with open(path, 'r', encoding='utf-8', errors='replace') as f:
                html = f.read()

            blocks = pattern_detect.detect_repeating_blocks(html)
            if blocks:
                rows = extract_list_fields(blocks[0]['items_html'], fields, api_key,
                                           provider=provider, model=model)
                for i, row in enumerate(rows):
                    row['_source_file'] = os.path.relpath(path, folder)
                    row['_item_index'] = i + 1
                records.extend(rows)
                log_fn(f'[AI 추출] {os.path.basename(path)}: 반복되는 항목 {blocks[0]["count"]}개 감지, '
                       f'{len(rows)}개 행 추출')
                continue

            record = extract_fields(html, fields, api_key, provider=provider, model=model)
            record['_source_file'] = os.path.relpath(path, folder)
            records.append(record)
            log_fn(f'[AI 추출] 완료: {os.path.basename(path)}')
        except Exception as e:
            log_fn(f'[AI 추출] 실패 ({os.path.basename(path)}): {e}')

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = f'extracted_{timestamp}'
    saved_paths = export_records(records, folder, base_name, list(export_formats))
    log_fn(f'[AI 추출] {len(records)}/{len(html_files)}개 페이지 추출 완료. '
           f'저장: {", ".join(saved_paths) if saved_paths else "없음"}')
    return records, saved_paths
